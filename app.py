"""
app.py — Notion Digest › Gmail
Maison Eric Kayser Asie Ltd
"""

import streamlit as st
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import os
from dotenv import load_dotenv
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

# ── Chargement des variables d'environnement / secrets ──────────────────────────────────
load_dotenv()

# Utiliser st.secrets si disponible (Streamlit Cloud), sinon os.getenv (local)
try:
    GMAIL_ADDRESS      = st.secrets['GMAIL_ADDRESS']
    GMAIL_APP_PASSWORD = st.secrets['GMAIL_APP_PASSWORD']
    NOTION_TOKEN       = st.secrets['NOTION_TOKEN']
    DATABASE_ID        = st.secrets['NOTION_DATABASE_ID']
    DEFAULT_EMAIL      = st.secrets.get('DEFAULT_RECIPIENT', '')
except (KeyError, AttributeError):
    GMAIL_ADDRESS      = os.getenv('GMAIL_ADDRESS')
    GMAIL_APP_PASSWORD = os.getenv('GMAIL_APP_PASSWORD')
    NOTION_TOKEN       = os.getenv('NOTION_TOKEN')
    DATABASE_ID        = os.getenv('NOTION_DATABASE_ID')
    DEFAULT_EMAIL      = os.getenv('DEFAULT_RECIPIENT', '')


# ── Helpers ───────────────────────────────────────────────────────────────────

def format_date(date_str):
    """Convertit une date ISO en format français."""
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.strftime('%d/%m/%Y')
    except:
        return date_str

def parse_date(date_str):
    """Parse une date depuis un string (Excel ou autre format)."""
    if not date_str or date_str == '':
        return None
    try:
        # Essayer format français d'abord
        dt = datetime.strptime(str(date_str), '%d/%m/%Y')
        return dt.strftime('%Y-%m-%d')
    except:
        try:
            # Essayer format ISO
            dt = datetime.strptime(str(date_str), '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except:
            return None

def get_task_by_title(tasks_map, title):
    """Trouve une tâche par son titre."""
    for task_id, task in tasks_map.items():
        if task['title'].strip() == title.strip():
            return task_id
    return None

def fetch_notion_tasks() -> tuple[list[dict], dict, str | None]:
    """Récupère les tâches depuis Notion et retourne aussi la map complète."""
    url = f"https://api.notion.com/v1/databases/{DATABASE_ID}/query"
    headers = {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json",
    }
    
    pages = []
    has_more = True
    next_cursor = None
    
    while has_more:
        payload = {"page_size": 100}
        if next_cursor:
            payload["start_cursor"] = next_cursor
            
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=10)
        except requests.exceptions.RequestException as e:
            return [], {}, f"Erreur réseau : {e}"

        if resp.status_code != 200:
            return [], {}, f"Erreur Notion {resp.status_code} : {resp.text}"
            
        data = resp.json()
        pages.extend(data.get('results', []))
        has_more = data.get('has_more', False)
        next_cursor = data.get('next_cursor')

    tasks_map = {}
    
    for page in pages:
        page_id = page['id']
        title = "Sans titre"
        parent_id = None
        status = None
        due_date = None
        reminder = None
        
        try:
            props = page['properties']
            
            for prop_value in props.values():
                if prop_value.get('type') == 'title':
                    texts = prop_value.get('title', [])
                    if texts:
                        title = texts[0]['plain_text']
                    break
                    
            if 'Parent task' in props and props['Parent task']['type'] == 'relation':
                rel_data = props['Parent task'].get('relation', [])
                if rel_data:
                    parent_id = rel_data[0]['id']
            
            if 'Statut de la tâche' in props and props['Statut de la tâche']['type'] == 'status':
                status_data = props['Statut de la tâche'].get('status')
                if status_data:
                    status = status_data.get('name')
            
            if 'Date d\'échéance' in props and props['Date d\'échéance']['type'] == 'date':
                date_data = props['Date d\'échéance'].get('date')
                if date_data:
                    due_date = date_data.get('start')
            
            if 'Rappel' in props and props['Rappel']['type'] == 'date':
                reminder_data = props['Rappel'].get('date')
                if reminder_data:
                    reminder = reminder_data.get('start')
                    
        except Exception:
            pass
            
        tasks_map[page_id] = {
            "title": title,
            "parent_id": parent_id,
            "subtasks": [],
            "status": status,
            "due_date": due_date,
            "reminder": reminder
        }

    root_tasks = []
    for pid, task in tasks_map.items():
        parent_id = task["parent_id"]
        if parent_id and parent_id in tasks_map:
            tasks_map[parent_id]["subtasks"].append(task)
        else:
            root_tasks.append(task)

    return root_tasks, tasks_map, None


def update_notion_task(task_id: str, status: str = None, due_date: str = None, reminder: str = None) -> tuple[bool, str]:
    """Met à jour une tâche dans Notion."""
    url = f"https://api.notion.com/v1/pages/{task_id}"
    headers = {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Notion-Version": "2022-06-28",
        "Content-Type": "application/json",
    }
    
    properties = {}
    
    # Mettre à jour le statut
    if status:
        properties["Statut de la tâche"] = {
            "status": {
                "name": status
            }
        }
    
    # Mettre à jour la date d'échéance
    if due_date:
        properties["Date d\'échéance"] = {
            "date": {
                "start": due_date
            }
        }
    
    # Mettre à jour le rappel
    if reminder:
        properties["Rappel"] = {
            "date": {
                "start": reminder
            }
        }
    
    if not properties:
        return True, "Aucune modification"
    
    payload = {"properties": properties}
    
    try:
        resp = requests.patch(url, headers=headers, json=payload, timeout=10)
        if resp.status_code in [200, 202]:
            return True, "Mise à jour réussie"
        else:
            return False, f"Erreur {resp.status_code}: {resp.text}"
    except Exception as e:
        return False, f"Erreur : {e}"


def import_from_excel(excel_file, tasks_map: dict) -> tuple[int, list]:
    """Importe et met à jour les tâches depuis un fichier Excel."""
    try:
        df = pd.read_excel(excel_file)
        
        updated = 0
        errors = []
        
        for idx, row in df.iterrows():
            task_title = row.get('Tâche', '').strip()
            new_status = row.get('Statut', '').strip() if pd.notna(row.get('Statut')) else None
            new_due_date = row.get('Date d\'échéance', '') if pd.notna(row.get('Date d\'échéance')) else None
            new_reminder = row.get('Rappel', '') if pd.notna(row.get('Rappel')) else None
            
            # Ignorer les lignes vides ou en-têtes
            if not task_title or task_title == 'Tâche':
                continue
            
            # Supprimer les préfixes des sous-tâches pour la recherche
            clean_title = task_title.replace('↳ ', '').strip()
            for i in range(5):
                clean_title = clean_title.replace('    ', '')
            
            # Trouver la tâche dans la map
            task_id = get_task_by_title(tasks_map, clean_title)
            
            if task_id:
                # Parser les dates
                due_date_parsed = parse_date(new_due_date) if new_due_date else None
                reminder_parsed = parse_date(new_reminder) if new_reminder else None
                
                # Mettre à jour
                ok, msg = update_notion_task(
                    task_id,
                    status=new_status if new_status else None,
                    due_date=due_date_parsed,
                    reminder=reminder_parsed
                )
                
                if ok:
                    updated += 1
                else:
                    errors.append(f"Ligne {idx + 2}: {task_title} - {msg}")
            else:
                errors.append(f"Ligne {idx + 2}: Tâche '{clean_title}' non trouvée dans Notion")
        
        return updated, errors
        
    except Exception as e:
        return 0, [f"Erreur de lecture du fichier: {str(e)}"]


def build_email_html(root_tasks: list[dict]) -> str:
    today = datetime.now().strftime('%d/%m/%Y')
    
    def render_task_html(task) -> str:
        status_badge = ""
        if task.get('status'):
            status_colors = {
                "Pas démarré": "#FF6B6B",
                "En cours": "#4ECDC4",
                "Terminé": "#95E1D3",
                "Archive": "#CCCCCC"
            }
            color = status_colors.get(task['status'], "#999999")
            status_badge = f" <span style='background-color:{color};color:white;padding:2px 6px;border-radius:3px;font-size:11px;'>{task['status']}</span>"
        
        due_date_str = ""
        if task.get('due_date'):
            formatted_date = format_date(task['due_date'])
            due_date_str = f" <span style='color:#E74C3C;'>📅 {formatted_date}</span>"
        
        reminder_str = ""
        if task.get('reminder'):
            formatted_reminder = format_date(task['reminder'])
            reminder_str = f" <span style='color:#F39C12;'>🔔 {formatted_reminder}</span>"
        
        html = f"<li style='margin:8px 0'><strong>{task['title']}</strong>{status_badge}{due_date_str}{reminder_str}"
        if task["subtasks"]:
            html += "<ul style='padding-left:20px; list-style-type:circle;margin-top:6px;'>"
            for sub in task["subtasks"]:
                html += render_task_html(sub)
            html += "</ul>"
        html += "</li>"
        return html

    if root_tasks:
        items = "".join(render_task_html(t) for t in root_tasks)
        task_block = f"<ul style='padding-left:20px'>{items}</ul>"
    else:
        task_block = "<p style='color:#888'>Aucune tâche trouvée dans votre base Notion.</p>"

    return f"""
    <html><body style='font-family:Arial,sans-serif;color:#333;max-width:600px;margin:auto'>
      <h2 style='color:#c8a96e'>📝 Récapitulatif Notion</h2>
      <p>Bonjour Alexandre,</p>
      <p>Voici vos tâches et sous-tâches du <strong>{today}</strong> :</p>
      {task_block}
      <br>
      <p style='font-size:12px;color:#888;'>
        <span style='background-color:#FF6B6B;color:white;padding:2px 6px;border-radius:3px;'>Pas démarré</span>
        <span style='background-color:#4ECDC4;color:white;padding:2px 6px;border-radius:3px;margin-left:4px;'>En cours</span>
        <span style='background-color:#95E1D3;color:white;padding:2px 6px;border-radius:3px;margin-left:4px;'>Terminé</span>
      </p>
      <p>Bonne journée !<br>
      <span style='color:#888;font-size:12px'>Notion Digest — Maison Eric Kayser Asie Ltd</span></p>
    </body></html>
    """

def flatten_tasks_for_text(root_tasks: list[dict], level=0) -> str:
    text = ""
    prefix = "• " if level == 0 else "    " * level + "- "
    for t in root_tasks:
        line = f"{prefix}{t['title']}"
        if t.get('status'):
            line += f" [{t['status']}]"
        if t.get('due_date'):
            line += f" 📅 {format_date(t['due_date'])}"
        if t.get('reminder'):
            line += f" 🔔 {format_date(t['reminder'])}"
        text += line + "\n"
        if t["subtasks"]:
            text += flatten_tasks_for_text(t["subtasks"], level + 1)
    return text

def export_to_excel(root_tasks: list[dict]) -> bytes:
    """Exporte les tâches dans un fichier Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tâches"
    
    header_fill = PatternFill(start_color="C8A96E", end_color="C8A96E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Tâche", "Niveau", "Statut", "Date d'échéance", "Rappel"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    def add_task_to_excel(task, level=0):
        ws.append([
            task['title'],
            "Sous-tâche" if level > 0 else "Principale",
            task.get('status', ''),
            format_date(task['due_date']) if task.get('due_date') else '',
            format_date(task['reminder']) if task.get('reminder') else ''
        ])
        
        row = ws.max_row
        for cell in ws[row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        if level > 0:
            ws[row][0].value = "    " * level + "↳ " + task['title']
        
        for subtask in task['subtasks']:
            add_task_to_excel(subtask, level + 1)
    
    for task in root_tasks:
        add_task_to_excel(task)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def send_digest(recipient: str) -> tuple[bool, str]:
    root_tasks, _, err = fetch_notion_tasks()
    if err:
        return False, err

    if not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD:
        return False, "GMAIL_ADDRESS ou GMAIL_APP_PASSWORD manquant"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"📝 Récapitulatif Notion — {datetime.now().strftime('%d/%m/%Y')}"
    msg["From"]    = GMAIL_ADDRESS
    msg["To"]      = recipient

    if root_tasks:
        text_body = "Bonjour Alexandre,\n\nVos tâches et sous-tâches :\n\n" + flatten_tasks_for_text(root_tasks)
    else:
        text_body = "Bonjour Alexandre,\n\nAucune tâche trouvée."
        
    msg.attach(MIMEText(text_body, "plain"))
    msg.attach(MIMEText(build_email_html(root_tasks), "html"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_ADDRESS, recipient, msg.as_string())
    except smtplib.SMTPAuthenticationError:
        return False, "Authentification Gmail échouée"
    except Exception as e:
        return False, f"Erreur SMTP : {e}"

    return True, f"✅ Email envoyé à {recipient} ({len(root_tasks)} tâche(s))."


# ── Interface Streamlit ───────────────────────────────────────────────────────

st.set_page_config(page_title="Notion Digest — Kayser", page_icon="📧")
st.title("📧 Notion Digest")
st.caption("Maison Eric Kayser Asie Ltd")
st.write("Générez et envoyez par e-mail un récapitulatif de vos tâches Notion.")
st.divider()

# Tab 1 : Prévisualisation
tab1, tab2, tab3 = st.tabs(["👁 Prévisualisation", "📤 Importer Excel", "🚀 Envoyer Email"])

with tab1:
    if st.button("Charger les tâches"):
        with st.spinner("Chargement depuis Notion..."):
            root_tasks, tasks_map, err = fetch_notion_tasks()
        if err:
            st.error(err)
            st.info("💡 Vérifiez que le bot Notion est bien invité sur la base de données.")
        elif not root_tasks:
            st.warning("Aucune tâche trouvée dans la base Notion.")
        else:
            st.success("Tâches trouvées :")
            
            def render_preview(tasks, level=0):
                for t in tasks:
                    indent = "&nbsp;" * (level * 8)
                    prefix = "•" if level == 0 else "◦"
                    
                    status_str = f" **[{t['status']}]**" if t.get('status') else ""
                    due_date_str = f" 📅 {format_date(t['due_date'])}" if t.get('due_date') else ""
                    reminder_str = f" 🔔 {format_date(t['reminder'])}" if t.get('reminder') else ""
                    
                    full_text = f"{indent} {prefix} **{t['title']}**{status_str}{due_date_str}{reminder_str}" if level == 0 else f"{indent} {prefix} {t['title']}{status_str}{due_date_str}{reminder_str}"
                    st.markdown(full_text)
                    if t["subtasks"]:
                        render_preview(t["subtasks"], level + 1)
                        
            render_preview(root_tasks)
            
            # Bouton export Excel
            excel_data = export_to_excel(root_tasks)
            st.download_button(
                label="📥 Télécharger en Excel",
                data=excel_data,
                file_name=f"Taches_Notion_{datetime.now().strftime('%d-%m-%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

with tab2:
    st.subheader("📤 Importer et mettre à jour depuis Excel")
    st.info("Modifiez le fichier Excel téléchargé et réuploadez-le pour mettre à jour Notion")
    
    uploaded_file = st.file_uploader("Choisir un fichier Excel", type=['xlsx'])
    
    if uploaded_file:
        if st.button("Importer et mettre à jour"):
            with st.spinner("Chargement de Notion..."):
                root_tasks, tasks_map, err = fetch_notion_tasks()
            
            if err:
                st.error(err)
            else:
                with st.spinner("Mise à jour en cours..."):
                    updated, errors = import_from_excel(uploaded_file, tasks_map)
                
                if updated > 0:
                    st.success(f"✅ {updated} tâche(s) mise(s) à jour")
                
                if errors:
                    st.warning(f"⚠️  {len(errors)} erreur(s) détectée(s):")
                    for error in errors:
                        st.write(f"• {error}")
                
                if updated == 0 and not errors:
                    st.info("Aucune tâche à mettre à jour")

with tab3:
    st.subheader("🚀 Envoyer le Digest par Email")
    
    target_email = st.text_input("Adresse e-mail de réception :", value=DEFAULT_EMAIL)
    
    if st.button("Envoyer le Digest", type="primary"):
        if not target_email:
            st.warning("Veuillez saisir une adresse e-mail.")
        else:
            with st.spinner("Extraction Notion et envoi via Gmail..."):
                ok, msg = send_digest(target_email)
            if ok:
                st.success(msg)
            else:
                st.error(msg)
